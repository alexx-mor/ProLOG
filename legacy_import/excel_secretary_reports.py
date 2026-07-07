"""Parser for manually maintained secretary Excel reports."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from legacy_import.models import ImportIssue, IssueSeverity, LegacyParseResult, LegacyReportRow
from legacy_import.normalizers import clean_text, parse_hours

EXPECTED_HEADERS = (
    "Дата",
    "Вид работ",
    "Должность",
    "Инженер",
    "Затраченное время",
    "Объект",
    "Нахождение",
)
MAX_HEADER_SCAN_ROWS = 20


def parse_secretary_report(path: Path) -> LegacyParseResult:
    """Read the old Excel workbook and return raw dated rows with parse issues."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[LegacyReportRow] = []
    issues: list[ImportIssue] = []
    try:
        for worksheet in workbook.worksheets:
            header = _find_header(worksheet)
            if header is None:
                issues.append(
                    ImportIssue(
                        severity=IssueSeverity.WARNING,
                        code="HEADER_NOT_FOUND",
                        message="Лист пропущен: не найден ожидаемый заголовок старого отчета",
                        sheet_name=worksheet.title,
                    )
                )
                continue
            header_row, start_column = header
            for row_number in range(header_row + 1, worksheet.max_row + 1):
                values = [
                    worksheet.cell(row=row_number, column=start_column + offset).value
                    for offset in range(len(EXPECTED_HEADERS))
                ]
                work_date = _parse_date(values[0])
                if work_date is None:
                    continue
                rows.append(
                    LegacyReportRow(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        work_date=work_date,
                        description=clean_text(values[1]),
                        position_text=clean_text(values[2]),
                        employee_text=clean_text(values[3]) or worksheet.title,
                        hours=parse_hours(values[4]),
                        object_text=clean_text(values[5]),
                        legacy_location_text=clean_text(values[6]),
                    )
                )
    finally:
        workbook.close()
    return LegacyParseResult(source_file=path, rows=rows, issues=issues)


def _find_header(worksheet) -> tuple[int, int] | None:
    max_row = min(worksheet.max_row, MAX_HEADER_SCAN_ROWS)
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row), start=1):
        cleaned = [clean_text(cell.value) for cell in row]
        for index, value in enumerate(cleaned):
            if value != EXPECTED_HEADERS[0]:
                continue
            candidate = tuple(cleaned[index : index + len(EXPECTED_HEADERS)])
            if candidate == EXPECTED_HEADERS:
                return row_number, index + 1
    return None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except (TypeError, ValueError):
            return None
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
