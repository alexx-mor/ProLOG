"""Формирование Excel-книги: отдельный лист для каждого сотрудника."""

from __future__ import annotations

import re
from collections import defaultdict
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from workbot.models import StoredReport


HEADERS = ["Дата", "Виды работ", "Затраченное время", "Объект", "Местонахождение"]
_FORBIDDEN_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def export_reports(
    reports: list[StoredReport],
    export_dir: Path,
    date_from: date | None = None,
    date_to: date | None = None,
) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    suffix = _period_suffix(date_from, date_to)
    path = _unique_path(export_dir / f"WorkBot_отчеты_{suffix}.xlsx")

    workbook = Workbook()
    workbook.remove(workbook.active)
    grouped: dict[str, list[StoredReport]] = defaultdict(list)
    for report in reports:
        grouped[report.employee_name].append(report)

    used_titles: set[str] = set()
    if not grouped:
        sheet = workbook.create_sheet("Нет данных")
        sheet["A1"] = "За выбранный период отчётов нет"
        sheet["A1"].font = Font(bold=True)
    else:
        for employee_name in sorted(grouped, key=str.casefold):
            sheet = workbook.create_sheet(_sheet_title(employee_name, used_titles))
            _fill_sheet(sheet, grouped[employee_name])

    workbook.save(path)
    return path


def _fill_sheet(sheet, reports: list[StoredReport]) -> None:
    sheet.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7C9D6")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    for report in sorted(reports, key=lambda item: (item.work_date, item.id)):
        sheet.append(
            [
                report.work_date,
                report.work_types,
                report.hours if report.hours > 0 else None,
                report.object_name,
                report.location,
            ]
        )
    for row in sheet.iter_rows(min_row=2):
        row[0].number_format = "DD.MM.YYYY"
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].number_format = "0.##"
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.vertical = "top"
            cell.alignment = alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    widths = [13, 48, 20, 34, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 26
    sheet.sheet_view.showGridLines = False


def _sheet_title(employee_name: str, used: set[str]) -> str:
    parts = employee_name.strip().split()
    surname = parts[0] if parts else "Сотрудник"
    base = _FORBIDDEN_SHEET_CHARS.sub("_", surname).strip("' ") or "Сотрудник"
    candidate = base[:31]
    if candidate.casefold() not in used:
        used.add(candidate.casefold())
        return candidate
    initials = "".join(f"{part[0]}." for part in parts[1:3] if part)
    candidate = f"{base} {initials}".strip()[:31]
    number = 2
    while candidate.casefold() in used:
        tail = f" {number}"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        number += 1
    used.add(candidate.casefold())
    return candidate


def _period_suffix(date_from: date | None, date_to: date | None) -> str:
    if date_from and date_to:
        return f"{date_from:%Y%m%d}-{date_to:%Y%m%d}"
    if date_from:
        return f"с_{date_from:%Y%m%d}"
    if date_to:
        return f"по_{date_to:%Y%m%d}"
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    index = 2
    while True:
        candidate = path.with_stem(f"{path.stem}_{index}")
        if not candidate.exists():
            return candidate
        index += 1
