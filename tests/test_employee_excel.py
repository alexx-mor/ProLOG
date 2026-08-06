"""Employee Excel import and export checks."""

from datetime import datetime

from openpyxl import Workbook, load_workbook

from excel_export import export_employees, import_employees
from models import Employee


class _EmployeeImportRecorder:
    def __init__(self) -> None:
        self.values: list[dict[str, str]] = []

    def import_employee(self, **values) -> int:
        self.values.append(values)
        return len(self.values)


def test_employee_hire_date_round_trips_through_excel(tmp_path) -> None:
    source = tmp_path / "employees-source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ФИО", "Должность", "Разряд", "Дата трудоустройства"])
    sheet.append(["Иванов Иван Иванович", "Слесарь", "2", datetime(2024, 3, 18)])
    workbook.save(source)

    recorder = _EmployeeImportRecorder()
    assert import_employees(source, recorder) == 1
    assert recorder.values[0]["hire_date"] == "2024-03-18"

    result = tmp_path / "employees-result.xlsx"
    export_employees(
        result,
        [Employee("Иванов Иван Иванович", "Слесарь", "2", hire_date="2024-03-18")],
    )
    exported = load_workbook(result, read_only=True, data_only=True)
    try:
        assert exported.active.cell(1, 6).value == "Дата трудоустройства"
        assert exported.active.cell(2, 6).value == "18.03.2024"
    finally:
        exported.close()
