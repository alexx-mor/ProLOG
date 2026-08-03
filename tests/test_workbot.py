from __future__ import annotations

import hashlib
import hmac
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from workbot.config import WorkBotConfig
from workbot.excel import export_reports
from workbot.legacy_parser import parse_legacy_reports
from workbot.models import ParsedReport
from workbot.parser import ReportParseError, parse_report
from workbot.service import WorkBotService
from workbot.storage import WorkBotStorage


REPORT_TEXT = """Дата: 30.07.2026
Виды работ: Монтаж шкафа
и проверка цепей
Затраченное время: 7,5 ч
Объект: Цех № 1
Местонахождение: Производство"""


class FakeClient:
    def __init__(self) -> None:
        self.messages: list[tuple[int, str]] = []
        self.files: list[Path] = []
        self.callbacks: list[tuple[str, dict]] = []
        self.last_message_kwargs: dict = {}

    def send_message(self, text: str, *, user_id=None, chat_id=None, **kwargs):
        self.messages.append((user_id or chat_id, text))
        self.last_message_kwargs = kwargs
        return {}

    def send_file(self, path: Path, caption: str, *, user_id: int):
        self.files.append(path)
        return {}

    def answer_callback(self, callback_id: str, **kwargs):
        self.callbacks.append((callback_id, kwargs))
        return {}


class ParserTests(unittest.TestCase):
    def test_parse_complete_report(self):
        report = parse_report(REPORT_TEXT)
        self.assertIsNotNone(report)
        self.assertEqual(report.work_date, date(2026, 7, 30))
        self.assertEqual(report.hours, 7.5)
        self.assertIn("проверка цепей", report.work_types)

    def test_normal_message_is_ignored(self):
        self.assertIsNone(parse_report("Доброе утро, начинаем работу."))

    def test_partial_report_has_clear_error(self):
        with self.assertRaisesRegex(ReportParseError, "объект"):
            parse_report("Дата: 30.07.2026\nВиды работ: Монтаж\nЧасы: 8")

    def test_parse_free_form_daily_report(self):
        reports = parse_legacy_reports(
            "29.07.26\nИванов И.И.\nГазетная 23\nСборка шкафа Жигалово\n8-17",
            date(2026, 7, 29),
            "Иван",
        )
        self.assertEqual(len(reports), 1)
        self.assertEqual(reports[0].employee_name, "Иванов И.И.")
        self.assertEqual(reports[0].report.hours, 8)
        self.assertEqual(reports[0].report.location, "Газетная 23")
        self.assertEqual(reports[0].report.object_name, "Жигалово")

    def test_parse_multiple_dates_and_employees(self):
        reports = parse_legacy_reports(
            "Петров П.П.\nСидоров С.С.\n27.07.26\n28.07.26\n8:00-19:00\nМонтаж оборудования",
            date(2026, 7, 28),
            "Автор",
        )
        self.assertEqual(len(reports), 4)
        self.assertEqual({item.report.work_date for item in reports}, {date(2026, 7, 27), date(2026, 7, 28)})
        self.assertTrue(all(item.report.hours == 10 for item in reports))

    def test_parse_compact_date_range(self):
        reports = parse_legacy_reports(
            "06-08.06.2026\nИванов И.И.\nКомандировка",
            date(2026, 6, 8),
            "Иван",
        )
        self.assertEqual(
            {item.report.work_date for item in reports},
            {date(2026, 6, 6), date(2026, 6, 7), date(2026, 6, 8)},
        )


class StorageAndExportTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.storage = WorkBotStorage(self.root / "workbot.sqlite3")
        self.storage.initialize()
        self.storage.upsert_user(10, "Иван", "Иванов", "ivanov")

    def tearDown(self):
        self.temp.cleanup()

    def test_deduplicates_source_message_and_exports_surname_sheet(self):
        received = datetime(2026, 7, 30, 12, 0)
        self.assertTrue(self.storage.record_message("m1", 50, 10, received, REPORT_TEXT))
        self.assertFalse(self.storage.record_message("m1", 50, 10, received, REPORT_TEXT))
        self.storage.bind_user(10, "Иванов Иван Иванович")
        parsed = ParsedReport(date(2026, 7, 30), "Монтаж", 8, "Цех", "Производство")
        self.storage.save_report("m1", 10, self.storage.employee_name_for(10, ""), parsed)

        path = export_reports(self.storage.reports(), self.root)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["Иванов"])
            values = list(workbook["Иванов"].values)
            self.assertEqual(values[0], ("Дата", "Виды работ", "Затраченное время", "Объект", "Местонахождение"))
            self.assertEqual(values[1][2], 8)
        finally:
            workbook.close()

    def test_verified_phone_cannot_belong_to_two_max_users(self):
        self.storage.upsert_user(11, "Петр", "Петров", "petrov")
        self.storage.save_verified_phone(10, "+79991234567")
        with self.assertRaisesRegex(ValueError, "другим пользователем"):
            self.storage.save_verified_phone(11, "+79991234567")


class ServiceTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        self.storage = WorkBotStorage(root / "workbot.sqlite3")
        self.storage.initialize()
        self.client = FakeClient()
        self.config = WorkBotConfig(
            token="test",
            owner_ids=frozenset({1}),
            database_path=root / "workbot.sqlite3",
            export_dir=root,
        )
        self.service = WorkBotService(self.config, self.storage, self.client)

    def tearDown(self):
        self.temp.cleanup()

    def test_group_report_is_collected_silently(self):
        self.service.handle_update(_update(10, REPORT_TEXT, chat_id=100, mid="report-1"))
        self.assertEqual(len(self.storage.reports()), 1)
        self.assertEqual(self.client.messages, [])

    def test_private_message_from_stranger_is_ignored(self):
        self.service.handle_update(_update(10, "/help", chat_id=None, mid="private-1"))
        self.assertEqual(self.client.messages, [])

    def test_owner_can_request_stats(self):
        self.service.handle_update(_update(1, "/stats", chat_id=None, mid="owner-1"))
        self.assertIn("Принято отчётов", self.client.messages[0][1])

    def test_user_can_request_verified_contact(self):
        self.service.handle_update(_update(10, "/register", chat_id=None, mid="register-1"))
        attachments = self.client.last_message_kwargs["attachments"]
        button = attachments[0]["payload"]["buttons"][0][0]
        self.assertEqual(button["type"], "request_contact")

    def test_signed_max_contact_is_saved(self):
        vcard = "BEGIN:VCARD\r\nTEL;TYPE=CELL:+7 (999) 123-45-67\r\nEND:VCARD"
        signature = hmac.new(b"test", vcard.encode(), hashlib.sha256).hexdigest()
        self.service.handle_update(_contact_update(10, vcard, signature))
        user = next(row for row in self.storage.users() if row["max_user_id"] == 10)
        self.assertEqual(user["verified_phone"], "+79991234567")
        self.assertIn("подтвержден", self.client.messages[-1][1])

    def test_unverified_contact_is_rejected(self):
        vcard = "BEGIN:VCARD\r\nTEL:+7 999 123-45-67\r\nEND:VCARD"
        self.service.handle_update(_contact_update(10, vcard, "wrong"))
        user = next(row for row in self.storage.users() if row["max_user_id"] == 10)
        self.assertEqual(user["verified_phone"], "")
        self.assertIn("не удалось подтвердить", self.client.messages[0][1])

    def test_owner_can_open_inline_menu(self):
        self.service.handle_update(_update(1, "/menu", chat_id=None, mid="menu-1"))
        attachments = self.client.last_message_kwargs["attachments"]
        self.assertEqual(attachments[0]["type"], "inline_keyboard")
        self.assertEqual(len(attachments[0]["payload"]["buttons"]), 4)

    def test_owner_callback_runs_action(self):
        self.service.handle_update(_callback_update(1, "menu:stats"))
        self.assertEqual(self.client.callbacks[0][0], "callback-1")
        self.assertIn("Принято отчётов", self.client.messages[0][1])

    def test_foreign_callback_is_denied(self):
        self.service.handle_update(_callback_update(10, "menu:stats"))
        self.assertEqual(self.client.callbacks[0][1]["notification"], "Доступ запрещён.")
        self.assertEqual(self.client.messages, [])

    def test_edited_report_replaces_previous_values(self):
        self.service.handle_update(_update(10, REPORT_TEXT, chat_id=100, mid="edited-1"))
        edited = _update(
            10,
            REPORT_TEXT.replace("7,5 ч", "8 ч").replace("Монтаж шкафа", "Наладка шкафа"),
            chat_id=100,
            mid="edited-1",
        )
        edited["update_type"] = "message_edited"
        self.service.handle_update(edited)
        reports = self.storage.reports()
        self.assertEqual(len(reports), 1)
        self.assertEqual(reports[0].hours, 8)
        self.assertIn("Наладка", reports[0].work_types)


def _update(sender_id: int, text: str, chat_id: int | None, mid: str):
    recipient = {"chat_type": "chat", "chat_id": chat_id} if chat_id is not None else {"chat_type": "dialog"}
    return {
        "update_type": "message_created",
        "timestamp": 1785412800000,
        "message": {
            "sender": {
                "user_id": sender_id,
                "first_name": "Иван",
                "last_name": "Иванов",
                "username": f"user{sender_id}",
                "is_bot": False,
            },
            "recipient": recipient,
            "timestamp": 1785412800000,
            "body": {"mid": mid, "text": text},
        },
    }


def _callback_update(sender_id: int, payload: str):
    return {
        "update_type": "message_callback",
        "timestamp": 1785412800000,
        "callback": {
            "timestamp": 1785412800000,
            "callback_id": "callback-1",
            "payload": payload,
            "user": {
                "user_id": sender_id,
                "first_name": "Иван",
                "last_name": "Иванов",
                "is_bot": False,
            },
        },
    }


def _contact_update(sender_id: int, vcf_info: str, signature: str):
    update = _update(sender_id, "", chat_id=None, mid="contact-1")
    update["message"]["body"]["attachments"] = [
        {
            "type": "contact",
            "payload": {
                "vcf_info": vcf_info,
                "hash": signature,
                "max_info": {"user_id": sender_id},
            },
        }
    ]
    return update


if __name__ == "__main__":
    unittest.main()
