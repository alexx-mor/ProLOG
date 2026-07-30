"""Excel import and export operations."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from constants import EXPORTS_DIR
from models import AppSettings, Employee, WorkLogEntry
from services import EmployeeService
from utils import safe_filename, unique_path


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def import_employees(path: Path, service: EmployeeService) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(max_row=1))]
        required = {"фио": None, "должность": None, "категория": None, "разряд": None}
        for index, header in enumerate(headers):
            if header in required:
                required[header] = index
        if required["фио"] is None:
            raise ValueError("В Excel-файле не найдена колонка 'ФИО'")

        imported = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            full_name = _value(row, required["фио"])
            if not full_name:
                continue
            service.import_employee(
                full_name=full_name,
                position=_value(row, required["должность"]),
                category=_value(row, required["категория"]) or _value(row, required["разряд"]),
            )
            imported += 1
        return imported
    finally:
        workbook.close()


def export_employees(path: Path, employees: list[Employee]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Сотрудники"
    _write_header(sheet, ["№", "ФИО", "Должность", "Разряд"])
    for row_index, employee in enumerate(employees, start=2):
        sheet.append(
            [
                row_index - 1,
                employee.full_name,
                employee.position,
                employee.category,
            ]
        )
    _autosize(sheet)
    workbook.save(path)
    return path


def export_work_report(entries: list[WorkLogEntry], work_date: date, settings: AppSettings | None = None) -> Path:
    path = unique_path(EXPORTS_DIR / f"Отчет_работ_{work_date.isoformat()}.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчет"
    _write_document_header(sheet, f"Отчет выполнения работ за {work_date.strftime('%d.%m.%Y')}", settings)
    _write_header(
        sheet,
        ["№", "Сотрудник", "Местонахождение", "Объект", "Изделие", "Вид работ", "Описание работ", "Часы", "Комментарий"],
        row=6,
    )
    for index, entry in enumerate(entries, start=1):
        sheet.append(_entry_row(index, entry))
    _autosize(sheet)
    workbook.save(path)
    return path


def export_shift_assignment(entries: list[WorkLogEntry], work_date: date, settings: AppSettings | None = None) -> Path:
    path = unique_path(EXPORTS_DIR / f"Сменное_задание_{work_date.isoformat()}.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Сменное задание"
    _write_document_header(sheet, f"Сменное задание на {work_date.strftime('%d.%m.%Y')}", settings)
    _write_header(sheet, ["№", "Сотрудник", "Местонахождение", "Объект", "Изделие", "Вид работ", "Задание", "Комментарий"], row=6)
    for index, entry in enumerate(entries, start=1):
        sheet.append(
            [
                index,
                entry.employee_name,
                entry.location_name,
                entry.object_name,
                entry.product_name,
                entry.work_type_name,
                entry.description,
                entry.comment,
            ]
        )
    _autosize(sheet)
    workbook.save(path)
    return path


def default_employee_export_path() -> Path:
    return unique_path(EXPORTS_DIR / f"{safe_filename('Сотрудники')}.xlsx")


def _value(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def _write_header(sheet, headers: list[str], row: int = 1) -> None:
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=title)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_document_header(sheet, title: str, settings: AppSettings | None) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.merge_cells("A1:I1")
    if not settings:
        return
    sheet["A2"] = f"Организация: {settings.organization_name}".strip()
    sheet["A3"] = f"Отдел: {settings.department_name}".strip()
    sheet["A4"] = f"Руководитель: {settings.leader_full_name}".strip()


def _entry_row(index: int, entry: WorkLogEntry) -> list[object]:
    return [
        index,
        entry.employee_name,
        entry.location_name,
        entry.object_name,
        entry.product_name,
        entry.work_type_name,
        entry.description,
        entry.hours,
        entry.comment,
    ]


def _autosize(sheet) -> None:
    for column_index, column_cells in enumerate(sheet.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width, 10), 45)
